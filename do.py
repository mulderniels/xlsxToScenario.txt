#made by n.mulder1@uu.nl 2016

#creates scenario-files based upon template.txt. Data for each scenario is defined in
#data.xlsx. template.txt has placeholders '{varName}' for each variable.
#converts float and int to string

#init
import os.path
from openpyxl import load_workbook

#do stuff
print('hallo')

data = load_workbook(filename = 'data.xlsx')
data = data.active #select the active sheet

scenarioN = data.max_row-1 #no of scenarios
varN = data.max_column-2 #no of variables

#put vars in dict
vars = []
for iVar in range(3, varN+2+1):
	vars.append(data.cell(row = 1, column = iVar).value)

#loop trough scenarios
for iScenario in range(2, scenarioN+1+1):
	fileResultName = 'output-'+str(data.cell(row = iScenario, column = 2).value)+'.txt'
	#use a different template for the first scenario of a set
	if os.path.isfile(fileResultName):
		fileTemplate = open('template.txt', 'r')
	else:
		fileTemplate = open('templateFirst.txt', 'r')
	#open or create the output file
	fileResult = open(fileResultName, 'a')
	for line in fileTemplate:
		for iVar in range(0, varN):
			content = data.cell(row = iScenario, column = iVar+3).value;
			line = line.replace('{'+vars[iVar]+'}', str(content))
		fileResult.write(line)
	fileResult.close();
	fileTemplate.close();

print('doei')